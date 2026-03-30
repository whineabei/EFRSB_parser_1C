from __future__ import annotations

import argparse
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional

import undetected_chromedriver as uc
from openpyxl import Workbook, load_workbook
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR / "307 шт. - для парсера.xlsx"
NEW_COLUMNS = [
    "Дата публикации",
    "судебный акт",
    "№ дела",
    "Арбитражный управляющий",
    "Адрес для корреспонденции",
    "Дата решения",
    "Место жительства",
    "статус",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Парсер сообщений о судебных актах из Fedresurs."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Путь к исходному Excel-файлу.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Путь к результирующему Excel-файлу.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Имя листа. Если не указано, берется первый лист.",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Запускать браузер без окна.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Ограничить число строк для тестового прогона.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=40,
        help="Таймаут ожидания элементов, сек.",
    )
    return parser.parse_args()


def run_command(command: List[str]) -> str:
    kwargs = {
        "capture_output": True,
        "text": True,
        "check": True,
    }
    if sys.platform.startswith("win"):
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        kwargs["startupinfo"] = startupinfo
        kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
    result = subprocess.run(command, **kwargs)
    return result.stdout.strip()


def detect_chrome_path() -> Path:
    candidates = [
        Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
        Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
        Path.home() / r"AppData\Local\Google\Chrome\Application\chrome.exe",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        "Chrome не найден. Установите Google Chrome или поправьте путь в detect_chrome_path()."
    )


def get_chrome_version(chrome_path: Path) -> str:
    powershell = [
        "powershell",
        "-NoProfile",
        "-Command",
        f"(Get-Item '{chrome_path}').VersionInfo.ProductVersion",
    ]
    version = run_command(powershell)
    if not re.match(r"^\d+\.\d+\.\d+\.\d+$", version):
        raise RuntimeError(f"Не удалось определить версию Chrome: {version!r}")
    return version


def normalize_text(value: Optional[str]) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def extract_only_date(value: str) -> str:
    match = re.search(r"\d{2}\.\d{2}\.\d{4}", value)
    return match.group(0) if match else normalize_text(value)


def extract_status(court_act: str) -> str:
    normalized = normalize_text(court_act).lower()
    if not normalized:
        return ""
    return "завершено" if "о завершении" in normalized else "не завершено"


def is_court_act_message(value: str) -> bool:
    return "сообщение о судебном акте" in normalize_text(value).lower()


def build_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_result.xlsx")


def wait_for_page(driver: WebDriver, timeout: int) -> None:
    WebDriverWait(driver, timeout).until(
        lambda current_driver: current_driver.execute_script("return document.readyState")
        == "complete"
    )


def safe_get(driver: WebDriver, url: str, timeout: int) -> None:
    last_error: Optional[Exception] = None
    for attempt in range(2):
        try:
            driver.get(url)
            wait_for_page(driver, timeout)
            WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            time.sleep(1.5)
            return
        except Exception as error:
            last_error = error
            message = str(error).lower()
            is_network_error = any(
                marker in message
                for marker in [
                    "err_connection_closed",
                    "err_connection_reset",
                    "err_connection_aborted",
                    "err_network_changed",
                    "err_timed_out",
                    "net::",
                ]
            )
            if attempt == 1 or not is_network_error:
                raise
            print(f"[retry] Повторяю открытие страницы после сетевой ошибки: {url}")
            time.sleep(2)

    if last_error:
        raise last_error


def parse_info_items(driver: WebDriver) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for item in driver.find_elements(By.CSS_SELECTOR, "div.info-item"):
        lines = [normalize_text(line) for line in item.text.splitlines() if normalize_text(line)]
        if len(lines) >= 2:
            result[lines[0]] = " ".join(lines[1:])
    return result


def extract_manager_name(driver: WebDriver) -> str:
    for element in driver.find_elements(By.CSS_SELECTOR, "div.name"):
        text = normalize_text(element.text)
        if text:
            return text
    return ""


def extract_correspondence_address(driver: WebDriver) -> str:
    page_text = normalize_text(driver.find_element(By.TAG_NAME, "body").text)
    match = re.search(
        r"Адрес для корреспонденции\s*(.*?)\s*(?:Эл\.\s*почта|Данные СРО АУ|Должник|Сообщение)",
        page_text,
        re.IGNORECASE,
    )
    return normalize_text(match.group(1)) if match else ""


def parse_message_details(driver: WebDriver, url: str, timeout: int) -> Dict[str, str]:
    safe_get(driver, url, timeout)
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "div.info-item"))
    )
    info_items = parse_info_items(driver)
    court_act = info_items.get("Судебный акт", "")

    return {
        "Тип сообщения": normalize_text(driver.find_element(By.TAG_NAME, "body").text).split(" Заказать выписку", 1)[0],
        "судебный акт": court_act,
        "№ дела": info_items.get("Дело", ""),
        "Арбитражный управляющий": extract_manager_name(driver),
        "Адрес для корреспонденции": extract_correspondence_address(driver),
        "Дата решения": info_items.get("Дата акта", ""),
        "Место жительства": info_items.get("Место жительства", ""),
        "статус": extract_status(court_act),
    }


def parse_card_messages(driver: WebDriver, url: str, timeout: int) -> List[Dict[str, str]]:
    safe_get(driver, url, timeout)
    messages: List[Dict[str, str]] = []
    seen_urls = set()

    for message_link in driver.find_elements(By.XPATH, "//a[contains(@href, 'bankruptmessages')]"):
        href = normalize_text(message_link.get_attribute("href"))
        if not href or href in seen_urls:
            continue

        row = message_link.find_element(By.XPATH, "./ancestor::tr[1]")
        row_text = normalize_text(row.text)
        link_text = normalize_text(message_link.text)
        if not is_court_act_message(link_text) and not is_court_act_message(row_text):
            continue

        cells = row.find_elements(By.XPATH, "./td")
        publication_date = extract_only_date(cells[0].text if cells else "")
        messages.append({"Дата публикации": publication_date, "message_url": href})
        seen_urls.add(href)

    return messages


def init_driver(
    chrome_path: Path,
    chrome_version: str,
    headless: bool,
) -> WebDriver:
    options = uc.ChromeOptions()
    options.add_argument("--window-size=1600,1000")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--lang=ru-RU")
    if headless:
        options.add_argument("--headless=new")

    print(
        f"[driver] Запускаю undetected_chromedriver для Chrome major "
        f"{chrome_version.split('.', 1)[0]}"
    )

    driver = uc.Chrome(
        options=options,
        version_main=int(chrome_version.split(".", 1)[0]),
        browser_executable_path=str(chrome_path),
        use_subprocess=True,
    )
    driver.set_page_load_timeout(120)
    return driver


def should_restart_driver(error: Exception) -> bool:
    message = str(error).lower()
    return any(
        marker in message
        for marker in [
            "no such window",
            "invalid session id",
            "target window already closed",
            "connection refused",
            "httpconnectionpool",
            "max retries exceeded",
            "web view not found",
            "err_connection_closed",
            "err_connection_reset",
            "err_connection_aborted",
            "err_network_changed",
            "net::",
        ]
    )


def restart_driver(
    driver: WebDriver,
    chrome_path: Path,
    chrome_version: str,
    headless: bool,
) -> WebDriver:
    try:
        driver.quit()
    except Exception:
        pass
    time.sleep(1)
    return init_driver(chrome_path, chrome_version, headless)


def read_input_rows(input_path: Path, sheet_name: Optional[str]) -> tuple[str, List[str], List[Dict[str, str]]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    headers = [normalize_text(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows: List[Dict[str, str]] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[index]: row[index] for index in range(len(headers))}
        rows.append(row_dict)

    return worksheet.title, headers, rows


def write_result(
    output_path: Path,
    sheet_name: str,
    headers: List[str],
    rows: List[Dict[str, str]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)

    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    workbook.save(output_path)


def process_rows(
    driver: WebDriver,
    input_rows: List[Dict[str, str]],
    source_headers: List[str],
    timeout: int,
    limit: Optional[int],
    chrome_path: Path,
    chrome_version: str,
    headless: bool,
) -> tuple[List[Dict[str, str]], WebDriver]:
    site_column = "Доп. рекв.: Сайт (банкротство) (К)"
    result_rows: List[Dict[str, str]] = []
    total = len(input_rows) if limit is None else min(len(input_rows), limit)
    current_driver = driver

    for index, source_row in enumerate(input_rows[:total], start=1):
        site_url = normalize_text(source_row.get(site_column))
        debtor_name = normalize_text(source_row.get("Ответчик (наименование) (СД)"))
        print(f"[{index}/{total}] {debtor_name or '<без имени>'}")

        base_row = {header: source_row.get(header, "") for header in source_headers}
        for column in NEW_COLUMNS:
            base_row.setdefault(column, "")

        if not site_url:
            print("  - ссылка отсутствует")
            continue

        try:
            messages = parse_card_messages(current_driver, site_url, timeout)
        except TimeoutException:
            print("  - карточка не загрузилась по таймауту")
            continue
        except Exception as error:
            if should_restart_driver(error):
                print("  - драйвер отвалился, перезапускаю браузер")
                current_driver = restart_driver(
                    current_driver, chrome_path, chrome_version, headless
                )
                try:
                    messages = parse_card_messages(current_driver, site_url, timeout)
                except Exception as retry_error:
                    print(f"  - ошибка открытия карточки после перезапуска: {retry_error}")
                    continue
            else:
                print(f"  - ошибка открытия карточки: {error}")
                continue

        if not messages:
            print("  - сообщения о судебном акте не найдены")
            continue

        print(f"  - найдено сообщений о судебном акте: {len(messages)}")
        for message in messages:
            output_row = dict(base_row)
            output_row["Дата публикации"] = message.get("Дата публикации", "")

            try:
                details = parse_message_details(
                    current_driver, message["message_url"], timeout
                )
                message_type = normalize_text(details.get("Тип сообщения", ""))
                court_act = normalize_text(details.get("судебный акт", ""))
                if "Сообщение о судебном акте" not in message_type or not court_act:
                    print("    * пропускаю запись: это не судебный акт или поле пустое")
                    continue

                output_row.update(details)
                print(
                    "    * "
                    f"{output_row['Дата публикации']} | "
                    f"{normalize_text(output_row['судебный акт'])[:80]}"
                )
            except TimeoutException:
                print(f"    * не удалось загрузить сообщение: {message['message_url']}")
            except NoSuchElementException:
                print(f"    * не найдены ожидаемые поля: {message['message_url']}")
            except Exception as error:
                if should_restart_driver(error):
                    print("    * драйвер отвалился на сообщении, перезапускаю и пробую еще раз")
                    current_driver = restart_driver(
                        current_driver, chrome_path, chrome_version, headless
                    )
                    try:
                        details = parse_message_details(
                            current_driver, message["message_url"], timeout
                        )
                        message_type = normalize_text(details.get("Тип сообщения", ""))
                        court_act = normalize_text(details.get("судебный акт", ""))
                        if "Сообщение о судебном акте" not in message_type or not court_act:
                            print("    * пропускаю запись после перезапуска: это не судебный акт или поле пустое")
                            continue

                        output_row.update(details)
                        print(
                            "    * "
                            f"{output_row['Дата публикации']} | "
                            f"{normalize_text(output_row['судебный акт'])[:80]}"
                        )
                    except Exception as retry_error:
                        print(f"    * ошибка сообщения после перезапуска: {retry_error}")
                else:
                    print(f"    * ошибка сообщения: {error}")

            result_rows.append(output_row)

    return result_rows, current_driver


def main() -> int:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve() if args.output else build_output_path(input_path)

    if not input_path.exists():
        raise FileNotFoundError(f"Файл не найден: {input_path}")

    chrome_path = detect_chrome_path()
    chrome_version = get_chrome_version(chrome_path)
    sheet_name, source_headers, input_rows = read_input_rows(input_path, args.sheet)
    output_headers = source_headers + [column for column in NEW_COLUMNS if column not in source_headers]

    driver = init_driver(chrome_path, chrome_version, args.headless)
    try:
        result_rows, driver = process_rows(
            driver=driver,
            input_rows=input_rows,
            source_headers=output_headers,
            timeout=args.timeout,
            limit=args.limit,
            chrome_path=chrome_path,
            chrome_version=chrome_version,
            headless=args.headless,
        )
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    write_result(output_path, sheet_name, output_headers, result_rows)
    print(f"[done] Результат сохранен: {output_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\n[stop] Прервано пользователем")
        raise SystemExit(130)
    except Exception as error:
        print(f"[error] {error}", file=sys.stderr)
        raise SystemExit(1)
